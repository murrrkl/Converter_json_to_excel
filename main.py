import pandas
import pandas as pd
import os
import openpyxl
from openpyxl.styles import Font
from openpyxl import Workbook

if os.path.isfile('FinalTable.xlsx'):
    os.remove('FinalTable.xlsx')

my_file = open("FinalTable.xlsx", "w+")
my_file.close()

pandas.read_json("gen.json").to_excel("output.xlsx")
path = "output.xlsx"

wb = Workbook()
sheet = wb.active

# Table header
c1 = sheet.cell(row = 1, column = 1)
c2 = sheet.cell(row = 1, column = 2)
c3 = sheet.cell(row = 1, column = 3)
c4 = sheet.cell(row = 1, column = 4)
c5 = sheet.cell(row = 1, column = 5)

c1.value = "Id"
c2.value = "Дата обращения"
c3.value = "Пол"
c4.value = "Признак"
c5.value = "Значение"

c1.font = Font(bold = True)
c2.font = Font(bold = True)
c3.font = Font(bold = True)
c4.font = Font(bold = True)
c5.font = Font(bold = True)
# Table header --end

df_len = len(pd.read_excel(path, "Sheet1"))

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

count_ib = 1
table_row = 2


def json_1():
    global df_len
    global wb
    global cell_obj
    global table_row

    for i in range(2, df_len + 2):

        cell_obj = sheet_obj.cell(row=i, column=3).value

        open_brackets = 0
        close_brackets = 0

        j = 1

        while j < len(cell_obj):

            if cell_obj[j] == '{':
                open_brackets += 1
                feature_name = ""
                j += 13

                while cell_obj[j] != "'":
                    feature_name += cell_obj[j]
                    j += 1

                while open_brackets != close_brackets:
                    time = ""
                    value = ""
                    if cell_obj[j] == '{':
                        open_brackets += 1
                        j += 8
                        while cell_obj[j] != ",":
                            if cell_obj != "'":
                                time += cell_obj[j]
                            j += 1

                        j += 12
                        while cell_obj[j] != "'":
                            value += cell_obj[j]
                            j += 1

                        c11 = sheet.cell(row=table_row, column=1)
                        c11.value = i - 1

                        c12 = sheet.cell(row=table_row, column=2)
                        class_name = sheet_obj.cell(row=i, column=2).value
                        c12.value = class_name

                        c13 = sheet.cell(row=table_row, column=3)
                        c13.value = feature_name

                        c14 = sheet.cell(row=table_row, column=4)
                        c14.value = time

                        c15 = sheet.cell(row=table_row, column=5)
                        c15.value = value

                        table_row += 1

                    elif cell_obj[j] == '}':
                        close_brackets += 1
                    j += 1

            j += 1

    wb.save("FinalTable.xlsx")

def json_2():
    global df_len
    global wb
    global cell_obj
    global table_row

    for i in range(2, df_len + 2):

        # Текущая история болезни
        cell_obj = sheet_obj.cell(row=i, column=13).value
        n = len(cell_obj) - 2
        cell_obj = cell_obj[1:n]
        open_brackets = 0
        close_brackets = 0

        j = cell_obj.find("id")
        j += 6
        id = ""
        date = ""
        pol = ""

        while cell_obj[j] != ",":
           id += cell_obj[j]
           j += 1

        j = cell_obj.find("value")
        j += 8

        while cell_obj[j] != ",":
           date += cell_obj[j]
           j += 1

        j = cell_obj.find("value", j)
        j += 8

        while cell_obj[j] != ",":
           pol += cell_obj[j]
           j += 1

        #print(id)
        #print(date)
        #print(pol)

        c11 = sheet.cell(row=table_row, column=1)
        c11.value = id

        c12 = sheet.cell(row=table_row, column=2)
        class_name = sheet_obj.cell(row=i, column=2).value
        c12.value = date

        c13 = sheet.cell(row=table_row, column=3)
        c13.value = pol


        while j < n-1:
            feature_str = ""
            if cell_obj[j] == "{":
                while  cell_obj[j] != "}":
                    feature_str += cell_obj[j]
                    j += 1

                #print(feature_str)

                feature_name = ""
                feature_value = ""
                k = feature_str.find("name")
                k += 7

                while feature_str[k] != ",":
                    feature_name += feature_str[k]
                    k += 1

                #print(feature_name, end=" ")

                k = feature_str.find("value")
                k += 8

                while feature_str[k] != "," and k != len(feature_str) - 1:
                    feature_value += feature_str[k]
                    k += 1
                if feature_value == "[":
                    feature_value = "нет"

                #print(feature_value)

                c14 = sheet.cell(row=table_row, column=4)
                c14.value = feature_name

                c15 = sheet.cell(row=table_row, column=5)
                c15.value = feature_value

                table_row += 1
            j += 1

    wb.save("FinalTable.xlsx")

json_2()
